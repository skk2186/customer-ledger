from datetime import date, datetime
from decimal import Decimal
from zipfile import ZipFile

from openpyxl import load_workbook

from customer_ledger.bookkeeping_service import (
    AllocationInput,
    PaymentInput,
    ShipmentInput,
    create_payment_workflow,
    create_shipment_with_initial_payment,
)
from customer_ledger.customer_service import archive_customer, create_customer
from customer_ledger.export_service import (
    CUSTOMER_HEADERS,
    SUMMARY_HEADERS,
    export_all_ledger_workbook,
    export_customer_workbook,
    export_summary_workbook,
)
from customer_ledger.extensions import db
from customer_ledger.models import Customer


def _customer(name: str) -> Customer:
    customer = create_customer(db.session, name)
    db.session.commit()
    return customer


def _shipment_input(customer_id: int, **changes) -> ShipmentInput:
    values = {
        "customer_id": customer_id,
        "shipment_date": date(2026, 5, 1),
        "total_amount_cents": 100_000,
        "freight_cents": 0,
        "unloading_fee_cents": 0,
        "returned_pallet_tonnage_hundredths": 0,
        "returned_pallet_amount_cents": 0,
        "issue_deduction_cents": 0,
        "area_hundredths": 0,
        "rounding_cents": 0,
        "description": "合成导出",
    }
    values.update(changes)
    return ShipmentInput(**values)


def _number(value) -> Decimal:
    return Decimal(str(value))


def _exported_date(value):
    return value.date() if isinstance(value, datetime) else value


def _assert_readback_is_plain_xlsx(workbook) -> None:
    assert workbook._external_links == []
    assert workbook.vba_archive is None
    assert all(sheet.freeze_panes is None for sheet in workbook.worksheets)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            assert all(cell.data_type != "f" for cell in row)


def _assert_ooxml_has_no_panes(path) -> None:
    with ZipFile(path) as archive:
        worksheet_xml = [
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        ]
        assert worksheet_xml
        assert all(b"<pane" not in archive.read(name) for name in worksheet_xml)


def test_customer_export_has_allocated_and_prepayment_rows(tmp_path, app):
    with app.app_context():
        customer = _customer("导出合成客户")
        create_shipment_with_initial_payment(
            db.session,
            _shipment_input(customer.id),
            120_000,
            "现金",
            "合成预收",
            "token-export-customer",
        )
        path = export_customer_workbook(db.session, customer.id, tmp_path / "customer.xlsx")

        workbook = load_workbook(path, data_only=False)
        sheet = workbook[customer.name]
        assert sheet.freeze_panes is None
        _assert_ooxml_has_no_panes(path)
        assert list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0] == CUSTOMER_HEADERS
        assert sheet["A1"].value == customer.name
        assert "A1:M1" in {str(item) for item in sheet.merged_cells.ranges}
        assert sheet["A3"].is_date
        assert _number(sheet["J3"].value) == Decimal("1000")
        assert _number(sheet["K3"].value) == Decimal("0")
        assert _number(sheet["J4"].value) == Decimal("200")
        assert _number(sheet["K4"].value) == Decimal("-200")
        assert str(sheet["M4"].value).startswith("预收款")
        assert _number(sheet["J5"].value) == Decimal("1200")
        assert _number(sheet["K5"].value) == Decimal("-200")
        assert sheet["A5"].value == "合计"
        assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)


def test_summary_and_all_ledger_exports_have_contract_order(tmp_path, app):
    with app.app_context():
        customer = _customer("汇总导出客户")
        create_shipment_with_initial_payment(
            db.session,
            _shipment_input(customer.id),
            120_000,
            "银行转账",
            "合成已收",
            "token-export-summary",
        )
        summary_path = export_summary_workbook(
            db.session, tmp_path / "summary.xlsx", date(2026, 5, 31)
        )
        all_path = export_all_ledger_workbook(db.session, tmp_path / "all.xlsx", date(2026, 5, 31))

        summary_sheet = load_workbook(summary_path, data_only=True)["客户汇总总表"]
        assert summary_sheet.freeze_panes is None
        _assert_ooxml_has_no_panes(summary_path)
        assert (
            list(summary_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            == SUMMARY_HEADERS
        )
        assert summary_sheet["A3"].value == "合计"
        assert _number(summary_sheet["J2"].value) == Decimal("1200")
        assert _number(summary_sheet["K2"].value) == Decimal("-200")
        assert _number(summary_sheet["J3"].value) == Decimal("1200")
        all_workbook = load_workbook(all_path, data_only=True)
        _assert_readback_is_plain_xlsx(all_workbook)
        _assert_ooxml_has_no_panes(all_path)
        assert all_workbook.sheetnames[0] == "客户汇总总表"
        assert customer.name in all_workbook.sheetnames


def test_complex_export_readback_covers_split_payments_negative_values_cutoff_and_archived(
    tmp_path, app
):
    with app.app_context():
        customer = _customer("复杂读回合成客户")
        first = create_shipment_with_initial_payment(
            db.session,
            _shipment_input(customer.id, shipment_date=date(2026, 5, 1)),
            0,
            "银行转账",
            "",
            "token-readback-first",
        )
        second = create_shipment_with_initial_payment(
            db.session,
            _shipment_input(
                customer.id, shipment_date=date(2026, 5, 2), total_amount_cents=200_000
            ),
            0,
            "银行转账",
            "",
            "token-readback-second",
        )
        create_shipment_with_initial_payment(
            db.session,
            _shipment_input(
                customer.id,
                shipment_date=date(2026, 5, 3),
                total_amount_cents=0,
                returned_pallet_amount_cents=20_000,
                rounding_cents=100,
            ),
            0,
            "银行转账",
            "",
            "token-readback-negative",
        )
        future = create_shipment_with_initial_payment(
            db.session,
            _shipment_input(
                customer.id,
                shipment_date=date(2026, 6, 15),
                total_amount_cents=100_000,
            ),
            0,
            "银行转账",
            "",
            "token-readback-future",
        )
        split_payment = create_payment_workflow(
            db.session,
            PaymentInput(
                customer_id=customer.id,
                payment_date=date(2026, 5, 4),
                amount_cents=150_000,
                payment_method="银行转账",
                description="合成拆分收款",
            ),
            "specified",
            [
                AllocationInput(shipment_id=first.id, amount_cents=60_000),
                AllocationInput(shipment_id=second.id, amount_cents=90_000),
            ],
            "token-readback-split-payment",
        )
        create_payment_workflow(
            db.session,
            PaymentInput(
                customer_id=customer.id,
                payment_date=date(2026, 5, 5),
                amount_cents=10_000,
                payment_method="现金",
                description="合成未分配收款",
            ),
            "none",
            [],
            "token-readback-unallocated",
        )
        create_payment_workflow(
            db.session,
            PaymentInput(
                customer_id=customer.id,
                payment_date=date(2026, 5, 6),
                amount_cents=30_000,
                payment_method="微信",
                description="合成未来分配",
            ),
            "specified",
            [AllocationInput(shipment_id=future.id, amount_cents=30_000)],
            "token-readback-future-payment",
        )
        assert len(split_payment.allocations) == 2
        archived = _customer("归档读回合成客户")
        create_shipment_with_initial_payment(
            db.session,
            _shipment_input(archived.id, shipment_date=date(2026, 5, 7)),
            0,
            "银行转账",
            "",
            "token-readback-archived",
        )
        archive_customer(db.session, archived)
        db.session.commit()

        cutoff = date(2026, 5, 31)
        path = export_all_ledger_workbook(db.session, tmp_path / "complex-cutoff.xlsx", cutoff)
        workbook = load_workbook(path, data_only=True)
        _assert_readback_is_plain_xlsx(workbook)
        _assert_ooxml_has_no_panes(path)
        assert workbook.sheetnames[0] == "客户汇总总表"
        assert archived.name in workbook.sheetnames
        sheet = workbook[customer.name]
        assert sheet["A1"].value == customer.name
        assert tuple(sheet.cell(2, index).value for index in range(1, 14)) == CUSTOMER_HEADERS
        total_row = sheet.max_row
        assert sheet.cell(total_row, 1).value == "合计"
        data_rows = list(range(3, total_row))
        assert all(_exported_date(sheet.cell(row, 1).value) <= cutoff for row in data_rows)
        date_rows = {
            _exported_date(sheet.cell(row, 1).value): row
            for row in data_rows
            if sheet.cell(row, 1).value
        }
        assert _number(sheet.cell(date_rows[date(2026, 5, 1)], 10).value) == Decimal("600")
        assert _number(sheet.cell(date_rows[date(2026, 5, 2)], 10).value) == Decimal("900")
        negative_row = date_rows[date(2026, 5, 3)]
        assert _number(sheet.cell(negative_row, 9).value) == Decimal("-200")
        assert _number(sheet.cell(negative_row, 11).value) == Decimal("-201")
        assert _number(sheet.cell(negative_row, 12).value) == Decimal("1")
        prepayment_rows = [
            row for row in data_rows if str(sheet.cell(row, 13).value).startswith("预收款")
        ]
        assert sorted(_number(sheet.cell(row, 10).value) for row in prepayment_rows) == [
            Decimal("100"),
            Decimal("300"),
        ]
        summary_sheet = workbook["客户汇总总表"]
        summary_row = next(
            row
            for row in range(2, summary_sheet.max_row)
            if summary_sheet.cell(row, 1).value == customer.name
        )
        assert _number(summary_sheet.cell(summary_row, 9).value) == Decimal("2800")
        assert _number(summary_sheet.cell(summary_row, 10).value) == Decimal("1900")
        assert _number(summary_sheet.cell(summary_row, 11).value) == Decimal("899")
        for column in range(2, 13):
            assert _number(sheet.cell(total_row, column).value) == _number(
                summary_sheet.cell(summary_row, column).value
            )

        later_path = export_all_ledger_workbook(
            db.session, tmp_path / "complex-after-future.xlsx", date(2026, 7, 1)
        )
        later_workbook = load_workbook(later_path, data_only=True)
        _assert_readback_is_plain_xlsx(later_workbook)
        _assert_ooxml_has_no_panes(later_path)
        later_sheet = later_workbook[customer.name]
        later_dates = {
            _exported_date(later_sheet.cell(row, 1).value): row
            for row in range(3, later_sheet.max_row)
            if later_sheet.cell(row, 1).value
        }
        assert date(2026, 6, 15) in later_dates
        assert _number(later_sheet.cell(later_dates[date(2026, 6, 15)], 10).value) == Decimal("300")
        assert not any(
            str(later_sheet.cell(row, 13).value).startswith("预收款")
            and _number(later_sheet.cell(row, 10).value) == Decimal("300")
            for row in range(3, later_sheet.max_row)
        )


def test_export_and_import_pages_are_user_facing_chinese(client, app, tmp_path):
    with app.app_context():
        customer = _customer("页面合成客户")
        app.config["EXPORTS_DIR"] = str(tmp_path / "exports")
        assert client.get(f"/customers/{customer.id}/export.xlsx").status_code == 200
    assert client.get("/exports/summary.xlsx").status_code == 200
    assert client.get("/exports/all-ledgers.xlsx").status_code == 200
    body = client.get("/imports/legacy").get_data(as_text=True)
    assert "旧账迁移" in body
    assert all(text not in body for text in ("Customer ledger", "Bookkeeping", "Read-only report"))

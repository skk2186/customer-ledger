"""Contract-driven Excel exports built from the calculation service."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select

from .calculation_service import (
    customer_ledger_rows,
    payment_unallocated_cents,
    summarize_customers,
)
from .models import Customer, Payment, PaymentAllocation
from .validation import validate_excel_sheet_name

SUMMARY_SHEET_NAME = "客户汇总总表"
CUSTOMER_HEADERS = (
    "日期",
    "总货款",
    "运费",
    "卸车费",
    "退板吨位",
    "退板金额",
    "问题扣费",
    "平方数",
    "应收款",
    "实收款",
    "欠款",
    "抹零",
    "付款方式",
)
SUMMARY_HEADERS = (
    "客户名称",
    "总货款",
    "运费",
    "卸车费",
    "退板吨位",
    "退板金额",
    "问题扣费",
    "平方数",
    "应收款",
    "实收款",
    "欠款",
    "抹零",
)
MONEY_COLUMNS = frozenset({2, 3, 4, 6, 7, 9, 10, 11, 12})
QUANTITY_COLUMNS = frozenset({5, 8})
_MONEY_FORMAT = "#,##0.00"
_QUANTITY_FORMAT = "#,##0.00"
_DATE_FORMAT = "yyyy-mm-dd"
_HEADER_FILL = PatternFill("solid", fgColor="DCE6F7")
_TITLE_FILL = PatternFill("solid", fgColor="2459A6")
_TOTAL_FILL = PatternFill("solid", fgColor="EEF3FA")
_BORDER = Border(bottom=Side(style="thin", color="D9E1F2"))


class ExportError(ValueError):
    """A user-correctable export contract or destination error."""


def _decimal_units(value: int, divisor: int = 100) -> Decimal:
    return Decimal(value) / Decimal(divisor)


def _destination(path: str | Path) -> Path:
    destination = Path(path)
    if destination.suffix.casefold() != ".xlsx":
        raise ExportError("导出文件必须使用 .xlsx 格式。")
    destination.parent.mkdir(parents=True, exist_ok=True)
    return destination


def _customer_name(customer: Customer) -> str:
    try:
        return validate_excel_sheet_name(customer.name)
    except ValueError as exc:
        raise ExportError(str(exc)) from exc


def _payment_method_text(session, shipment_id: int, as_of: date | None = None) -> str:
    statement = (
        select(PaymentAllocation, Payment)
        .join(Payment, Payment.id == PaymentAllocation.payment_id)
        .where(
            PaymentAllocation.shipment_id == shipment_id,
            PaymentAllocation.active.is_(True),
            Payment.active.is_(True),
        )
        .order_by(Payment.payment_date.asc(), Payment.id.asc(), PaymentAllocation.id.asc())
    )
    if as_of is not None:
        statement = statement.where(Payment.payment_date <= as_of)
    values = []
    for _allocation, payment in session.execute(statement).all():
        text = payment.payment_method.strip()
        if payment.description.strip():
            text = f"{text} / {payment.description.strip()[:40]}"
        if text and text not in values:
            values.append(text)
    return "、".join(values)


def _shipment_row(session, ledger_row, as_of: date | None = None) -> list[object]:
    shipment = ledger_row.shipment
    calculation = ledger_row.calculation
    return [
        shipment.shipment_date,
        _decimal_units(shipment.total_amount_cents),
        _decimal_units(shipment.freight_cents),
        _decimal_units(shipment.unloading_fee_cents),
        _decimal_units(shipment.returned_pallet_tonnage_hundredths),
        _decimal_units(shipment.returned_pallet_amount_cents),
        _decimal_units(shipment.issue_deduction_cents),
        _decimal_units(shipment.area_hundredths),
        _decimal_units(calculation.receivable_cents),
        _decimal_units(calculation.received_cents),
        _decimal_units(calculation.balance_cents),
        _decimal_units(shipment.rounding_cents),
        _payment_method_text(session, shipment.id, as_of),
    ]


def _prepayment_row(payment: Payment, unallocated_cents: int) -> list[object]:
    method = f"预收款 / {payment.payment_method.strip()}"
    if payment.description.strip():
        method = f"{method} / {payment.description.strip()[:40]}"
    return [
        payment.payment_date,
        Decimal(0),
        Decimal(0),
        Decimal(0),
        Decimal(0),
        Decimal(0),
        Decimal(0),
        Decimal(0),
        Decimal(0),
        _decimal_units(unallocated_cents),
        _decimal_units(-unallocated_cents),
        Decimal(0),
        method,
    ]


def _customer_rows(
    session, customer: Customer, as_of: date | None = None
) -> list[list[object]]:
    rows = [
        _shipment_row(session, row, as_of)
        for row in customer_ledger_rows(session, customer.id, as_of=as_of)
        if row.shipment.active
    ]
    payment_statement = select(Payment).where(
        Payment.customer_id == customer.id, Payment.active.is_(True)
    )
    if as_of is not None:
        payment_statement = payment_statement.where(Payment.payment_date <= as_of)
    payments = session.scalars(
        payment_statement.order_by(Payment.payment_date.asc(), Payment.id.asc())
    ).all()
    for payment in payments:
        unallocated = payment_unallocated_cents(session, payment, as_of=as_of)
        if unallocated > 0:
            rows.append(_prepayment_row(payment, unallocated))
    return rows


def _configure_title_row(sheet, title: str) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    cell = sheet.cell(1, 1, title)
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.fill = _TITLE_FILL
    cell.alignment = Alignment(horizontal="center")
    for column in range(1, 14):
        sheet.cell(1, column).fill = _TITLE_FILL


def _configure_header(sheet, row_number: int, headers: tuple[str, ...]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row_number, column, header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER


def _write_customer_sheet(
    session, workbook: Workbook, customer: Customer, as_of: date | None = None
) -> None:
    sheet = workbook.create_sheet(_customer_name(customer))
    _configure_title_row(sheet, customer.name)
    _configure_header(sheet, 2, CUSTOMER_HEADERS)
    rows = _customer_rows(session, customer, as_of=as_of)
    for row_number, values in enumerate(rows, start=3):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            if column == 1:
                cell.number_format = _DATE_FORMAT
            elif column in MONEY_COLUMNS:
                cell.number_format = _MONEY_FORMAT
            elif column in QUANTITY_COLUMNS:
                cell.number_format = _QUANTITY_FORMAT
    _write_customer_total(sheet, len(rows) + 3)
    sheet.freeze_panes = "A3"
    _finish_sheet(sheet, CUSTOMER_HEADERS)


def _write_customer_total(sheet, row_number: int) -> None:
    sheet.cell(row_number, 1, "合计")
    for column in range(2, 13):
        values = [sheet.cell(index, column).value for index in range(3, row_number)]
        total = sum((value or Decimal(0) for value in values), Decimal(0))
        cell = sheet.cell(row_number, column, total)
        cell.number_format = _MONEY_FORMAT if column in MONEY_COLUMNS else _QUANTITY_FORMAT
    sheet.cell(row_number, 13, "—")
    for column in range(1, 14):
        cell = sheet.cell(row_number, column)
        cell.font = Font(bold=True)
        cell.fill = _TOTAL_FILL
        cell.border = _BORDER


def _write_summary_sheet(session, workbook: Workbook, as_of: date) -> None:
    sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    _configure_header(sheet, 1, SUMMARY_HEADERS)
    summary_rows, grand_total = summarize_customers(session, as_of)
    for row_number, item in enumerate(summary_rows, start=2):
        summary = item.summary
        values = [
            item.customer.name,
            _decimal_units(summary.total_goods_cents),
            _decimal_units(summary.total_freight_cents),
            _decimal_units(summary.total_unloading_fee_cents),
            _decimal_units(summary.total_tonnage_hundredths),
            _decimal_units(summary.total_returned_pallet_cents),
            _decimal_units(summary.total_issue_deduction_cents),
            _decimal_units(summary.total_area_hundredths),
            _decimal_units(summary.total_receivable_cents),
            _decimal_units(summary.total_received_cents),
            _decimal_units(summary.net_balance_cents),
            _decimal_units(summary.total_rounding_cents),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            if column > 1:
                cell.number_format = (
                    _MONEY_FORMAT if column not in QUANTITY_COLUMNS else _QUANTITY_FORMAT
                )
    total_row = len(summary_rows) + 2
    sheet.cell(total_row, 1, "合计")
    totals = [
        grand_total.total_goods_cents,
        grand_total.total_freight_cents,
        grand_total.total_unloading_fee_cents,
        grand_total.total_tonnage_hundredths,
        grand_total.total_returned_pallet_cents,
        grand_total.total_issue_deduction_cents,
        grand_total.total_area_hundredths,
        grand_total.total_receivable_cents,
        grand_total.total_received_cents,
        grand_total.net_balance_cents,
        grand_total.total_rounding_cents,
    ]
    for column, value in enumerate(totals, start=2):
        cell = sheet.cell(total_row, column, _decimal_units(value))
        cell.number_format = _MONEY_FORMAT if column not in QUANTITY_COLUMNS else _QUANTITY_FORMAT
    for column in range(1, 13):
        cell = sheet.cell(total_row, column)
        cell.font = Font(bold=True)
        cell.fill = _TOTAL_FILL
        cell.border = _BORDER
    sheet.freeze_panes = "A2"
    _finish_sheet(sheet, SUMMARY_HEADERS)


def _finish_sheet(sheet, headers: tuple[str, ...]) -> None:
    widths = {
        column: max(len(str(headers[column - 1])) + 2, 12)
        for column in range(1, len(headers) + 1)
    }
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = min(max(widths[cell.column], len(str(cell.value)) + 2), 28)
    for column, width in widths.items():
        sheet.column_dimensions[chr(64 + column)].width = width
    sheet.sheet_view.showGridLines = False


def export_customer_workbook(session, customer_id: int, destination: str | Path) -> Path:
    customer = session.get(Customer, customer_id)
    if customer is None:
        raise ExportError("客户不存在。")
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_customer_sheet(session, workbook, customer)
    output = _destination(destination)
    workbook.save(output)
    return output


def export_summary_workbook(session, destination: str | Path, as_of: date | None = None) -> Path:
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_summary_sheet(session, workbook, as_of or date.max)
    output = _destination(destination)
    workbook.save(output)
    return output


def export_all_ledger_workbook(session, destination: str | Path, as_of: date | None = None) -> Path:
    workbook = Workbook()
    del workbook[workbook.sheetnames[0]]
    cutoff = as_of or date.max
    _write_summary_sheet(session, workbook, cutoff)
    customers = session.scalars(select(Customer).order_by(Customer.name.asc())).all()
    for customer in customers:
        _write_customer_sheet(session, workbook, customer, as_of=cutoff)
    output = _destination(destination)
    workbook.save(output)
    return output

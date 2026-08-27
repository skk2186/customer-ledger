from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from customer_ledger import create_app
from customer_ledger.bookkeeping_service import (
    AllocationInput,
    PaymentInput,
    ShipmentInput,
    create_payment_workflow,
    create_shipment_with_initial_payment,
)
from customer_ledger.customer_service import create_customer
from customer_ledger.export_service import (
    CUSTOMER_COLUMN_WIDTHS,
    CUSTOMER_HEADERS,
    SUMMARY_COLUMN_WIDTHS,
    export_customer_workbook,
    export_summary_workbook,
)
from customer_ledger.extensions import db
from customer_ledger.settings_service import (
    load_export_directory,
    save_export_directory,
)


def _customer(name: str):
    customer = create_customer(db.session, name)
    db.session.commit()
    return SimpleNamespace(id=customer.id, name=customer.name)


def _shipment_input(customer_id: int, **changes) -> ShipmentInput:
    values = {
        "customer_id": customer_id,
        "shipment_date": date(2026, 8, 27),
        "total_amount_cents": 100_000,
        "freight_cents": 0,
        "unloading_fee_cents": 0,
        "returned_pallet_tonnage_hundredths": 0,
        "returned_pallet_amount_cents": 0,
        "issue_deduction_cents": 0,
        "area_hundredths": 0,
        "rounding_cents": 0,
        "description": "合成用户体验测试",
    }
    values.update(changes)
    return ShipmentInput(**values)


def test_export_settings_json_is_utf8_atomic_and_corrupt_settings_fall_back(tmp_path):
    settings_path = tmp_path / "settings.json"
    default_directory = tmp_path / "默认导出"
    custom_directory = tmp_path / "自定义导出"

    save_export_directory(settings_path, custom_directory)
    assert settings_path.read_text(encoding="utf-8").endswith("\n")
    assert json.loads(settings_path.read_text(encoding="utf-8")) == {
        "export_directory": str(custom_directory.resolve())
    }
    assert load_export_directory(settings_path, default_directory) == custom_directory.resolve()
    assert not list(tmp_path.glob(".settings.json.*.tmp"))

    settings_path.write_text("不是合法 JSON", encoding="utf-8")
    assert load_export_directory(settings_path, default_directory) == default_directory.resolve()

    settings_path.write_text(
        json.dumps(
            {
                "export_directory": str(custom_directory),
                "DATABASE_PATH": str(tmp_path / "should-not-be-read.db"),
                "BACKUP_DIR": str(tmp_path / "should-not-be-read-backups"),
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    assert load_export_directory(settings_path, default_directory) == custom_directory.resolve()


def test_export_settings_route_persists_custom_directory_and_restores_default(
    client, app, tmp_path
):
    custom_directory = tmp_path / "custom-export"
    response = client.post(
        "/settings/export",
        data={"action": "save", "export_directory": str(custom_directory)},
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert "当前导出文件夹" in response.get_data(as_text=True)
    assert app.config["EXPORTS_DIR"] == str(custom_directory.resolve())
    assert Path(app.config["SETTINGS_PATH"]).is_file()

    app2 = create_app(
        {
            "TESTING": True,
            "SQLALCHEMY_DATABASE_URI": app.config["SQLALCHEMY_DATABASE_URI"],
            "BACKUP_DIR": app.config["BACKUP_DIR"],
            "SETTINGS_PATH": app.config["SETTINGS_PATH"],
            "DEFAULT_EXPORTS_DIR": app.config["DEFAULT_EXPORTS_DIR"],
        }
    )
    assert app2.config["EXPORTS_DIR"] == str(custom_directory.resolve())

    response = client.post(
        "/settings/export",
        data={"action": "reset", "export_directory": str(custom_directory)},
        follow_redirects=True,
    )
    assert response.status_code == 200
    assert app.config["EXPORTS_DIR"] == app.config["DEFAULT_EXPORTS_DIR"]


def test_export_settings_reports_file_path_instead_of_silent_fallback(client, tmp_path):
    existing_file = tmp_path / "not-a-folder"
    existing_file.write_text("synthetic", encoding="utf-8")
    response = client.post(
        "/settings/export",
        data={"action": "save", "export_directory": str(existing_file)},
    )
    assert response.status_code == 200
    assert "不是文件夹" in response.get_data(as_text=True)


def test_exports_use_chinese_unique_names_and_success_feedback(client, app, tmp_path):
    with app.app_context():
        app.config["EXPORTS_DIR"] = str(tmp_path / "exports")
        customer = _customer("罗桥孙总")

    first = client.get(f"/customers/{customer.id}/export.xlsx?feedback=1")
    second = client.get(f"/customers/{customer.id}/export.xlsx?feedback=1")
    assert first.status_code == 200
    assert second.status_code == 200
    assert "导出成功" in first.get_data(as_text=True)
    assert "已保存到" in first.get_data(as_text=True)
    assert "打开导出文件夹" in first.get_data(as_text=True)
    exported = sorted(Path(app.config["EXPORTS_DIR"]).glob("*.xlsx"))
    assert [path.name for path in exported] == [
        f"罗桥孙总_客户账目_{date.today():%Y%m%d}.xlsx",
        f"罗桥孙总_客户账目_{date.today():%Y%m%d}_2.xlsx",
    ]

    summary = client.get("/exports/summary.xlsx?as_of=2026-08-27&feedback=1")
    assert summary.status_code == 200
    assert Path(app.config["EXPORTS_DIR"], "客户汇总总表_20260827.xlsx").is_file()


def test_export_workbooks_use_fixed_widths_and_large_amount_formats(tmp_path, app):
    with app.app_context():
        customer = _customer("大金额排版客户")
        create_shipment_with_initial_payment(
            db.session,
            _shipment_input(
                customer.id,
                total_amount_cents=12_345_678_999,
                freight_cents=2_345_678_999,
                area_hundredths=9_876_543_210,
            ),
            9_876_543_210,
            "银行转账",
            "合成大金额",
            "token-large-export",
        )
        customer_path = export_customer_workbook(db.session, customer.id, tmp_path / "large.xlsx")
        summary_path = export_summary_workbook(
            db.session, tmp_path / "summary.xlsx", date(2026, 8, 27)
        )

    workbook = load_workbook(customer_path, data_only=False)
    sheet = workbook[customer.name]
    assert tuple(sheet.cell(2, index).value for index in range(1, 14)) == CUSTOMER_HEADERS
    assert {column: sheet.column_dimensions[column].width for column in CUSTOMER_COLUMN_WIDTHS} == {
        column: float(width) for column, width in CUSTOMER_COLUMN_WIDTHS.items()
    }
    assert sheet["B3"].number_format == "#,##0.00"
    assert sheet["I3"].number_format == "#,##0.00"
    assert sheet["J3"].number_format == "#,##0.00"
    assert sheet["K3"].number_format == "#,##0.00"
    assert sheet["B4"].number_format == "#,##0.00"
    assert sheet["B3"].alignment.horizontal == "right"
    assert sheet["A3"].alignment.horizontal == "center"
    assert sheet["M3"].alignment.wrap_text is True
    assert sheet["B3"].data_type != "f"
    assert workbook._external_links == []
    assert workbook.vba_archive is None

    summary_sheet = load_workbook(summary_path, data_only=False)["客户汇总总表"]
    assert {
        column: summary_sheet.column_dimensions[column].width
        for column in SUMMARY_COLUMN_WIDTHS
    } == {column: float(width) for column, width in SUMMARY_COLUMN_WIDTHS.items()}
    assert summary_sheet["J2"].number_format == "#,##0.00"
    assert summary_sheet["K2"].number_format == "#,##0.00"


def test_export_http_download_keeps_chinese_filename(client, app, tmp_path):
    with app.app_context():
        app.config["EXPORTS_DIR"] = str(tmp_path / "exports")
        customer = _customer("下载文件名客户")
    response = client.get(f"/customers/{customer.id}/export.xlsx")
    assert response.status_code == 200
    assert "attachment" in response.headers["Content-Disposition"]
    assert "客户账目" in response.headers["Content-Disposition"] or "%E5%AE%A2" in response.headers[
        "Content-Disposition"
    ]


def test_all_major_pages_have_explicit_parent_and_home_navigation(client, app):
    with app.app_context():
        customer = _customer("导航合成客户")

    pages = [
        "/customers",
        "/customers/new",
        f"/customers/{customer.id}/edit",
        f"/customers/{customer.id}/ledger",
        f"/shipments/new?customer_id={customer.id}",
        "/shipments/new",
        f"/payments/new?customer_id={customer.id}",
        "/payments/new",
        "/retail/new",
        "/summary",
        "/backups",
        "/audit",
        "/imports/legacy",
        "/settings/export",
    ]
    for page in pages:
        response = client.get(page)
        body = response.get_data(as_text=True)
        assert response.status_code == 200, page
        assert "返回上一级" in body, page
        assert "返回首页" in body, page
        assert "history.back" not in body, page

    customer_form = client.get("/customers/new").get_data(as_text=True)
    assert "/customers" in customer_form
    customer_ledger = client.get(f"/customers/{customer.id}/ledger").get_data(as_text=True)
    assert "/customers" in customer_ledger
    shipment_form = client.get(f"/shipments/new?customer_id={customer.id}").get_data(as_text=True)
    assert f"/customers/{customer.id}/ledger" in shipment_form


def test_payment_correspondence_ui_uses_dates_and_plain_chinese(client, app):
    with app.app_context():
        customer = _customer("对应关系合成客户")
        shipment = create_shipment_with_initial_payment(
            db.session,
            _shipment_input(customer.id),
            0,
            "银行转账",
            "",
            "token-correspondence-shipment",
        )
        create_payment_workflow(
            db.session,
            PaymentInput(
                customer_id=customer.id,
                payment_date=date(2026, 8, 27),
                amount_cents=100_000,
                payment_method="微信",
                description="合成对应",
            ),
            "specified",
            [AllocationInput(shipment.id, 40_000)],
            "token-correspondence-payment",
        )

    body = client.get(f"/customers/{customer.id}/ledger").get_data(as_text=True)
    assert "收款金额" in body
    assert "已用于货款" in body
    assert "剩余预收" in body
    assert "用于某笔货款" in body
    assert "查看对应货款" in body
    assert "2026-08-27 发货" in body
    assert "取消这笔对应" in body
    assert "查看分配" not in body
    assert ">分配<" not in body
    assert "发货 #" not in body
